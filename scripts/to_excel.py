#!/usr/bin/env python3
"""
to_excel.py — ProtoSync measurement log → formatted Excel workbook
Usage:  python3 to_excel.py protosync_YYYYMMDD_HHMMSS.csv [output.xlsx]

Produces a workbook with four sheets:
  1. Raw Data    — all rows, auto-filtered, colour-coded jitter
  2. MIDI Stats  — per-second MIDI clock statistics summary
  3. GPIO Stats  — per-second GPIO clock statistics summary
  4. Charts      — jitter stddev over time (MIDI vs GPIO overlay)
"""

import sys
import os
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Colour palette (matches LCD UI colours where sensible) ───────────────────
C_HEADER_BG   = "1C1C38"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_MIDI_BG     = "0D3D2D"   # dark teal
C_GPIO_BG     = "1A2540"   # dark blue
C_GOOD        = "3CE650"   # green   < 500 µs
C_WARN        = "FFC800"   # amber   < 2000 µs
C_ERR         = "FF3232"   # red     ≥ 2000 µs
C_LIGHT_GRAY  = "F4F4F4"
C_ALT_ROW     = "EAEFF7"

JITTER_GOOD   = 500.0
JITTER_WARN   = 2000.0


def jitter_colour(value: float) -> str:
    av = abs(value)
    if av < JITTER_GOOD:
        return C_GOOD
    if av < JITTER_WARN:
        return C_WARN
    return C_ERR


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────────
def write_raw_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Raw Data"

    cols = [
        ("elapsed_s",          "Elapsed (s)",        10),
        ("source",             "Source",              7),
        ("bpm",                "BPM",                 8),
        ("interval_mean_us",   "Mean (µs)",          12),
        ("interval_min_us",    "Min (µs)",            11),
        ("interval_max_us",    "Max (µs)",            11),
        ("interval_stddev_us", "σ Jitter (µs)",      14),
        ("jitter_last_us",     "Last Jitter (µs)",   15),
        ("jitter_peak_us",     "Peak Jitter (µs)",   15),
        ("sync_error_us",      "Sync Error (µs)",    15),
        ("sample_count",       "Samples",            10),
    ]

    # Header row
    for col_idx, (field, label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=10)
        cell.fill      = header_fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        is_midi = row.source == "MIDI"
        row_bg  = C_MIDI_BG if (is_midi and row_idx % 2 == 0) else \
                  C_GPIO_BG if (not is_midi and row_idx % 2 == 0) else \
                  C_LIGHT_GRAY if row_idx % 2 == 0 else "FFFFFF"

        for col_idx, (field, label, _) in enumerate(cols, start=1):
            val  = getattr(row, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right")
            cell.border    = thin_border()

            # Colour-code jitter columns
            if field in ("interval_stddev_us", "jitter_last_us",
                         "jitter_peak_us", "sync_error_us") and row.sample_count > 0:
                jcol = jitter_colour(float(val))
                cell.fill = PatternFill("solid", fgColor=jcol, start_color=jcol)
            elif field == "source":
                src_col = "00B4D8" if is_midi else "90E0EF"
                cell.fill = PatternFill("solid", fgColor=src_col, start_color=src_col)
                cell.font = Font(name="Arial", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = PatternFill("solid", fgColor=row_bg, start_color=row_bg)

            # Number formats
            if field in ("elapsed_s", "bpm"):
                cell.number_format = "0.000"
            elif field != "source" and field != "sample_count":
                cell.number_format = "0.00"


# ── Sheet 2/3: Stats summary per source ──────────────────────────────────────
def write_stats_sheet(wb: Workbook, df: pd.DataFrame, source: str, title: str):
    ws = wb.create_sheet(title=title)

    sub = df[df["source"] == source].copy()
    if sub.empty:
        ws["A1"] = f"No {source} data found."
        return

    metric_cols = [
        "interval_mean_us",
        "interval_min_us",
        "interval_max_us",
        "interval_stddev_us",
        "jitter_peak_us",
        "sync_error_us",
    ]
    labels = {
        "interval_mean_us":   "Mean Interval (µs)",
        "interval_min_us":    "Min Interval (µs)",
        "interval_max_us":    "Max Interval (µs)",
        "interval_stddev_us": "Stddev / 1σ Jitter (µs)",
        "jitter_peak_us":     "Peak |Jitter| (µs)",
        "sync_error_us":      "Link Sync Error (µs)",
    }

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = f"{source} Clock — Timing Statistics Summary"
    title_cell.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=12)
    title_cell.fill      = header_fill(C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Sub-header
    stat_names = ["Metric", "Min", "Max", "Mean", "Median", "Stddev"]
    for col_idx, s in enumerate(stat_names, start=1):
        cell = ws.cell(row=2, column=col_idx, value=s)
        cell.font      = Font(bold=True, name="Arial", size=10)
        cell.fill      = header_fill("3A3A6A")
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border()

    col_widths = {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14}
    set_col_widths(ws, col_widths)

    # One row per metric
    for row_idx, col in enumerate(metric_cols, start=3):
        data = sub[col].dropna()
        stats = [
            labels[col],
            round(float(data.min()),    2),
            round(float(data.max()),    2),
            round(float(data.mean()),   2),
            round(float(data.median()), 2),
            round(float(data.std()),    2),
        ]
        alt = row_idx % 2 == 0
        for col_idx, val in enumerate(stats, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.border    = thin_border()
            bg = C_ALT_ROW if alt else "FFFFFF"
            cell.fill      = PatternFill("solid", fgColor=bg, start_color=bg)
            if col_idx > 1:
                cell.alignment    = Alignment(horizontal="right")
                cell.number_format = "0.00"
                # Colour-code jitter stddev mean
                if col == "interval_stddev_us" and col_idx == 4:
                    jcol = jitter_colour(float(val))
                    cell.fill = PatternFill("solid", fgColor=jcol, start_color=jcol)

    # BPM range note
    note_row = len(metric_cols) + 4
    ws.cell(row=note_row, column=1,
            value=f"BPM range: {sub['bpm'].min():.1f} – {sub['bpm'].max():.1f}")
    ws.cell(row=note_row, column=1).font = Font(italic=True, name="Arial", size=9)

    ws.cell(row=note_row + 1, column=1,
            value=f"Samples logged: {int(sub['sample_count'].max())}")
    ws.cell(row=note_row + 1, column=1).font = Font(italic=True, name="Arial", size=9)

    ws.cell(row=note_row + 2, column=1,
            value="Colour:  GREEN < 500 µs  |  AMBER < 2000 µs  |  RED ≥ 2000 µs")
    ws.cell(row=note_row + 2, column=1).font = Font(italic=True, name="Arial", size=9, color="555555")

    ws.freeze_panes = "A3"


# ── Sheet 4: Charts ───────────────────────────────────────────────────────────
def write_chart_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet(title="Charts")

    midi = df[df["source"] == "MIDI"].reset_index(drop=True)
    gpio = df[df["source"] == "GPIO"].reset_index(drop=True)

    # Write compact data tables for the chart references
    # MIDI: cols A-C  (elapsed_s, interval_stddev_us, sync_error_us)
    # GPIO: cols E-F  (elapsed_s, interval_stddev_us)
    ws["A1"] = "elapsed_s";            ws["B1"] = "MIDI σ (µs)"; ws["C1"] = "MIDI SyncErr (µs)"
    ws["E1"] = "elapsed_s";            ws["F1"] = "GPIO σ (µs)"

    for i, row in midi.iterrows():
        ws.cell(row=i + 2, column=1, value=round(row["elapsed_s"],           3))
        ws.cell(row=i + 2, column=2, value=round(row["interval_stddev_us"],  2))
        ws.cell(row=i + 2, column=3, value=round(row["sync_error_us"],       2))

    for i, row in gpio.iterrows():
        ws.cell(row=i + 2, column=5, value=round(row["elapsed_s"],           3))
        ws.cell(row=i + 2, column=6, value=round(row["interval_stddev_us"],  2))

    n_midi = len(midi)
    n_gpio = len(gpio)

    # ── Chart 1: Jitter σ over time (MIDI vs GPIO) ────────────────────────────
    chart1 = LineChart()
    chart1.title   = "Jitter σ over Time — MIDI vs GPIO"
    chart1.style   = 10
    chart1.y_axis.title = "1σ Jitter (µs)"
    chart1.x_axis.title = "Elapsed (s)"
    chart1.height  = 12
    chart1.width   = 22
    chart1.y_axis.scaling.min = 0

    midi_sigma = Reference(ws, min_col=2, min_row=1, max_row=n_midi + 1)
    gpio_sigma = Reference(ws, min_col=6, min_row=1, max_row=n_gpio + 1)

    chart1.add_data(midi_sigma, titles_from_data=True)
    chart1.add_data(gpio_sigma, titles_from_data=True)

    chart1.series[0].graphicalProperties.line.solidFill = "00B4D8"   # teal = MIDI
    chart1.series[0].graphicalProperties.line.width     = 15000
    chart1.series[1].graphicalProperties.line.solidFill = "F4A261"   # orange = GPIO
    chart1.series[1].graphicalProperties.line.width     = 15000

    ws.add_chart(chart1, "H1")

    # ── Chart 2: MIDI Link Sync Error ─────────────────────────────────────────
    if n_midi > 0:
        chart2 = LineChart()
        chart2.title   = "MIDI → Link Sync Error over Time"
        chart2.style   = 10
        chart2.y_axis.title = "Sync Error (µs)"
        chart2.x_axis.title = "Elapsed (s)"
        chart2.height  = 12
        chart2.width   = 22

        sync_data = Reference(ws, min_col=3, min_row=1, max_row=n_midi + 1)
        chart2.add_data(sync_data, titles_from_data=True)
        chart2.series[0].graphicalProperties.line.solidFill = "E63946"
        chart2.series[0].graphicalProperties.line.width     = 15000

        ws.add_chart(chart2, "H24")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 to_excel.py <protosync_*.csv> [output.xlsx]")
        sys.exit(1)

    csv_path = sys.argv[1]
    if not os.path.exists(csv_path):
        print(f"Error: file not found: {csv_path}")
        sys.exit(1)

    xlsx_path = sys.argv[2] if len(sys.argv) >= 3 \
                else os.path.splitext(csv_path)[0] + ".xlsx"

    print(f"Reading {csv_path} …")
    df = pd.read_csv(csv_path)

    print(f"  {len(df)} rows  |  "
          f"MIDI: {(df['source']=='MIDI').sum()}  "
          f"GPIO: {(df['source']=='GPIO').sum()}")

    wb = Workbook()

    print("Writing Raw Data sheet …")
    write_raw_sheet(wb, df)

    print("Writing MIDI Stats sheet …")
    write_stats_sheet(wb, df, "MIDI", "MIDI Stats")

    print("Writing GPIO Stats sheet …")
    write_stats_sheet(wb, df, "GPIO", "GPIO Stats")

    print("Writing Charts sheet …")
    write_chart_sheet(wb, df)

    wb.save(xlsx_path)
    print(f"Saved → {xlsx_path}")


if __name__ == "__main__":
    main()